from flask import Flask, render_template, request, redirect, url_for
import os
from datetime import datetime
from PIL import Image
import requests
import time
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024

UPLOAD_FOLDER = "fotos"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ☁️ SIRV
CLIENT_ID = "8ARvedBbjChk50zgogeCBPtR5sN"
CLIENT_SECRET = "egU7oS4olUnvB0VrVXrEnhRquiVCRAJ86ZLDsFNsDq0z632VxpVJDdAJQPQ+fIrxlIT5QeWKjnsNE4c5zLSlHw=="
SIRV_DOMAIN = "https://gusdovi.sirv.com"


# ----------------------------
# SUBIR ARCHIVO A SIRV
# ----------------------------
def subir_archivo(file_path, extension, carpeta):
    try:
        auth = requests.post(
            "https://api.sirv.com/v2/token",
            json={"clientId": CLIENT_ID, "clientSecret": CLIENT_SECRET}
        )

        if auth.status_code != 200:
            print("Error auth:", auth.text)
            return None

        token = auth.json().get("token")

        filename = f"{datetime.now().timestamp()}.{extension}"
        upload_url = f"https://api.sirv.com/v2/files/upload?filename=/{carpeta}/{filename}"

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream"
        }

        with open(file_path, "rb") as f:
            response = requests.post(upload_url, headers=headers, data=f.read())

        if response.status_code != 200:
            print("Error subida:", response.text)
            return None

        return f"{SIRV_DOMAIN}/{carpeta}/{filename}"

    except Exception as e:
        print("Error:", e)
        return None


# ----------------------------
# DESCARGAR EXCEL EXISTENTE
# ----------------------------
def descargar_excel():
    url = f"{SIRV_DOMAIN}/reportes/maestro.xlsx"
    local_file = "maestro.xlsx"

    try:
        r = requests.get(url)
        if r.status_code == 200:
            with open(local_file, "wb") as f:
                f.write(r.content)
            return local_file
        else:
            return None
    except:
        return None


# ----------------------------
# SUBIR IMAGEN
# ----------------------------
def subir_imagen(file):
    try:
        img = Image.open(file)

        if img.mode != "RGB":
            img = img.convert("RGB")

        img.thumbnail((600, 600))

        temp = os.path.join(UPLOAD_FOLDER, "temp.jpg")
        img.save(temp, optimize=True, quality=50)

        url = subir_archivo(temp, "jpg", "obra")

        os.remove(temp)

        return url if url else "error_imagen"

    except:
        return "error_imagen"


# ----------------------------
# FORMULARIO
# ----------------------------
@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":

        obra = request.form["obra"]
        nombre = request.form["nombre"]
        actividad = request.form["actividad"]
        etapa = request.form["etapa"]
        ubicacion = request.form["ubicacion"]
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

        # 📷 imagen
        foto = request.files.get("foto")
        url_imagen = "sin_foto"

        if foto and foto.filename != "":
            url_imagen = subir_imagen(foto)

        # 📊 Excel acumulativo
        archivo = descargar_excel()

        if archivo:
            wb = load_workbook(archivo)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"
            ws.append(["Obra", "Obrero", "Ubicación", "Actividad", "Etapa", "Fecha", "Foto"])

        ws.append([obra, nombre, ubicacion, actividad, etapa, fecha, url_imagen])

        wb.save("maestro.xlsx")

        time.sleep(0.5)

        subir_archivo("maestro.xlsx", "xlsx", "reportes")

        os.remove("maestro.xlsx")

        return redirect(url_for("success"))

    return render_template("form.html")


# ----------------------------
# SUCCESS
# ----------------------------
@app.route("/success")
def success():
    return render_template("success.html")


# ----------------------------
# RUN
# ----------------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)